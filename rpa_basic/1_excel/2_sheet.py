from openpyxl import Workbook
wb = Workbook()
# ws = wb.active()
ws = wb.create_sheet() # Create the new sheet as a basic name

ws.title = "MySheet" # Change the sheet name
ws.sheet_properties.tabcolor = "ff66ff" # Tab color can be change when the value as RGB type is inserted

# Sheet, MySheet, YourSheet
ws1 = wb.create_sheet("YourSheet") # Sheet created with the defined name
ws2 = wb.create_sheet("NewSheet", 2) # Sheet created in the 2nd index

new_ws = wb["NewSheet"] # Sheet can be accessed with the Dict type

print(wb.sheetnames) # Check the names of all sheets

# Sheet Copy
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("Sample.xlsx")