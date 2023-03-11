import openpyxl

save_path = './StartCoding_Python_Automation/status_buying_A.xlsx'

# load a Excel file
wb = openpyxl.load_workbook('./StartCoding_Python_Automation/status_buying_A.xlsx')

# choose the activated sheet
ws = wb.active

# add a data (1)
ws['A1'] = 'Date'
ws['B1'] = 'Name of Products'
ws['C1'] = 'Price'
ws['D1'] = 'Quantity'
ws['E1'] = 'Total'

# add a data (2)
ws.cell(row=2, column=1, value='2030-01-01')
ws.cell(row=2, column=2, value='Gaming Mouse')
ws.cell(row=2, column=3, value=50000)
ws.cell(row=2, column=4, value=30)
ws.cell(row=2, column=5, value='=C2*D2')

# add a data (3)
ws.append(['2030-01-03', 'Mechanical Keyboard', 120000, 15, '=C3*D3' ])

# data modification
ws['C2'] = 40000
ws['D2'] = 40

# delete data
del ws['A3']

# save Excel
wb.save(save_path)