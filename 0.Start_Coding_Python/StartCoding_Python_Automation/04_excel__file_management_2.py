import openpyxl

# create a new Excel file
wb = openpyxl.Workbook()

# create a new sheet
ws = wb.create_sheet('2030.01')

# Print all of the names of the worksheets
print(wb.sheetnames)

# Delete the sheet
del wb['Sheet']

# save Excel
wb.save('./StartCoding_Python_Automation/status_buying_A.xlsx')