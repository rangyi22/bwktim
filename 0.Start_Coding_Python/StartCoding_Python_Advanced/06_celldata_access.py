import openpyxl

save_path = 'StartCoding_Python_Advanced/smartstore.xlsx'

# loading the current Excel file
wb = openpyxl.load_workbook(save_path, data_only=True)

# Choose the datasheet
ws = wb['data']

# 01. Load all cell data
# Case for knowing the number of column and rows
# for x in range(1, 9 + 1):
#     for y in range(1, 5 + 1):
#         print(ws.cell(row=x, column=y).value, end=" ")
#     print()

# Case for not knowking the number of column and rows
# for x in range(1, ws.max_row + 1):
#     for y in range(1, ws.max_column + 1):
#         print(ws.cell(row=x, column=y).value, end=" ")
#     print()

# Load all rows
# for row in ws.iter_rows():
#     print(row)

# Load from the 2nd rows
# for row in ws.iter_rows(min_row=2):
#     print(row)

# Load from the 2nd row to the 5th row
# for row in ws.iter_rows(min_row=2, max_row=5):
#     print(row)

# Load both 2 ~ 4 rows and 2 ~ 3 columns
for row in ws.iter_rows(min_row=2, max_row=4, min_col=2, max_col=4):
    for cell in row:
        print(cell.value, end=" ")
    print()