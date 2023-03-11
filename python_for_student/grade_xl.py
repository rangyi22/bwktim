#grade_xl.py
#> pip install openpyxl (in Windows Powershell)
import openpyxl as opx
# Excel 파일 불러오기
Mypath = 'D:/Python/data/'
fxl = Mypath+'grade.xlsx'; fxl1 = Mypath+'grade1.xlsx'
wb = opx.load_workbook(fxl) # Workbook
# Active sheet 불러오기
ws = wb.active # Worksheet # ws = wb["Sheet1"]
Number_of_keys = ws.max_column
Number_of_data = ws.max_row-1 # header (label) row 제외
New_column_no = Number_of_keys + 1
for i in range(0,Number_of_data+1):
   if i==0:  # 제0 (header) row에 대해
     sum = 'Sum'
   else: # data를 포함한 제1 row에서 A마지막 row까지
     kor = ws.cell(row=i+1, column=2).value
     mat = ws.cell(row=i+1, column=3).value
     phy = ws.cell(row=i+1 column=4).value
     sum = kor + mat + phy
   ws.cell(row=i+1, column=New_column_no).value = sum
# 수정된 workbook을 다른 Excel 파일로 저장
wb.save(fxl1)
wb.close()
