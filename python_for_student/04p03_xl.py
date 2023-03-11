#04p03_xl.py
import openpyxl as opx
wb = opx.load_workbook('py04p03_1.xlsx')
ws1 = wb['Sheet1']
ws2 = wb['Sheet2']
M1 = ws1.max_row;  N1 = ws1.max_column
M2 = ws2.max_row;  N2 = ws2.max_column
for m in range(1,  +1):
   for n in range(1, N2+ ):
      tmp = ws2.cell(row=  , column=n).value
      ws1.cell(row=m, column=     ).value = tmp   
wb.save('py04p03_2.xlsx');  wb.close()
