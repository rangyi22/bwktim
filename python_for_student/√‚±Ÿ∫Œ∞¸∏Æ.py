#출근부관리.py
# > python 출근부관리.py 홍길동 출: 홍길동(출석부).xlsx에 날짜/출근시각 
# > python 출근부관리.py 홍길동 퇴: 홍길동(출석부).xlsx에 퇴근시각 기록
# > python 출근부관리.py 홍길동 : 홍길동(출석부).xlsx 내용을 다 보여줘
import sys
from datetime import datetime
import openpyxl as opx
from openpyxl.workbook.workbook import *

def show_출석부(): # 출석부 내용을 다 print한다. 
   for m in range(1, row_max+1):
      date =  ws1.cell(row=m, column=1).value
      if type(date) is not str: # (제1 열) 날짜가 str형이 아니면
        date = date.strftime('%Y-%m-%d') # str형으로 바꿔서 
      print(date[0:10], end=' ') # print하고 한 칸을 띠고 이어서       
      for n in range(2, col_max+1): # 제2 열부터 끝 열까지 print해
         print(ws1.cell(row=m, column=n).value, end=' ')
      print()   

def 출근시간기록(): # 출근날짜와 시각을 기록한다.
   row = row_max
   day = ws1.cell(row, 1).value
   if type(day) is not str: #Excel에 의해 datetime 객체로 바뀐 경우
     day = day.strftime('%Y-%m-%d') # str형으로 변환
   if day is not today: row += 1
   # 마지막행의 날짜가 오늘(today)이 아닌 경우에만 다음 행에다
   ws1.cell(row, 1).value=today # 현재날짜를 제1 열에 
   ws1.cell(row, 2).value=now.strftime('%H:%M') #현재시각을 제2 열에

if len(sys.argv) > 1:
  # python 명령어뒤에 실행파일명외 이름 등 입력인자가 한 개 이상인 경우에만
  name = sys.argv[1]
  fname = '출근부(' + name + ').xlsx'
  try: wb = opx.load_workbook(fname) #기존 파일이 있으면 불러들여
  except: # 없으면 새 파일을 만든다.
    wb = Workbook() # ~.workbook module 내에 정의된 Workbook class
    wb.save(fname)
  sheet_names = wb.sheetnames
  ws1 = wb[sheet_names[0]] # 제0 worksheet
  row_max = ws1.max_row
  col_max = ws1.max_column  
  if row_max <= 1: # 첫 row
    ws1.append({'A': '날짜', 'B': '출근', 'C': '퇴근'})
    wb.save(fname)
  now = datetime.now() # 현재 시각
  today = now.strftime('%Y-%m-%d') # 현재 날짜
  if len(sys.argv)<3: # 입력인자가 (이 파일명과) 출근자 이름뿐인 경우
    show_출석부() 
  else:
    출퇴=sys.argv[2] # 이 파일의 이름, 출근자 이름에 이어 세 번째 입력인자
    if 출퇴[0] == '출':
      출근시간기록()                                        
    else: # 퇴근시간을 제3열에 기록한다.
      ws1.cell(row_max,3).value = now.strftime('%H:%M')
  wb.save(fname)
  wb.close()
